"""
Gera relatório XLSX de uma validação (achados + sugestões de correção).
Abas: Resumo, Achados, Sugestões
"""
from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from app.models.correction import CorrectionSuggestion
from app.models.efd_file import EfdFile
from app.models.fiscal_period import FiscalPeriod
from app.models.validation import ValidationFinding, ValidationRun

# Paleta de cores por severidade
SEVERITY_COLORS = {
    "critico":               "FFCCCC",  # vermelho claro
    "alerta":                "FFF2CC",  # amarelo claro
    "divergencia_monetaria": "CCE5FF",  # azul claro
    "observacao":            "F0F0F0",  # cinza
}

SEVERITY_LABELS = {
    "critico":               "Crítico",
    "alerta":                "Alerta",
    "divergencia_monetaria": "Divergência Monetária",
    "observacao":            "Observação",
}

FINDING_TYPE_LABELS = {
    "divergencia_monetaria":      "Divergência monetária",
    "ausencia_efd":               "Ausência no EFD",
    "ausencia_referencia":        "Sem referência",
    "sem_referencia_revisada":    "Referência não revisada",
    "registro_obrigatorio_ausente": "Registro obrigatório ausente",
    "codigo_invalido":            "Código inválido",
}

STATUS_LABELS = {
    "open":         "Aberto",
    "acknowledged": "Ciente",
    "resolved":     "Resolvido",
}


def _header_style(ws, row: int, cols: int, title: str, fill_hex: str = "1F3864") -> None:
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _col_header(ws, row: int, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="2E5090")
    font = Font(bold=True, color="FFFFFF", size=10)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit(ws, min_w: int = 10, max_w: int = 50) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_w))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, max_len + 2)


def _fmt_dec(v) -> str:
    if v is None:
        return ""
    try:
        return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)


def generate_xlsx(
    run: ValidationRun,
    findings: list[ValidationFinding],
    suggestions: list[CorrectionSuggestion],
    period: FiscalPeriod,
    efd_file: EfdFile,
) -> bytes:
    wb = openpyxl.Workbook()

    _build_summary_sheet(wb.active, run, findings, period, efd_file)
    wb.active.title = "Resumo"

    findings_ws = wb.create_sheet("Achados")
    _build_findings_sheet(findings_ws, findings)

    if suggestions:
        sug_ws = wb.create_sheet("Sugestões")
        _build_suggestions_sheet(sug_ws, suggestions)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(
    ws,
    run: ValidationRun,
    findings: list[ValidationFinding],
    period: FiscalPeriod,
    efd_file: EfdFile,
) -> None:
    ws.freeze_panes = "A1"

    # Título
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "FiscalCheck EFD — Relatório de Conferência"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadados
    meta = [
        ("Competência",        f"{period.month:02d}/{period.year}"),
        ("Arquivo EFD",        efd_file.original_filename),
        ("CNPJ (EFD)",         efd_file.efd_cnpj or "—"),
        ("Empresa (EFD)",      efd_file.efd_company_name or "—"),
        ("Período EFD",        f"{efd_file.efd_start_date or '?'} → {efd_file.efd_end_date or '?'}"),
        ("Data da conferência",run.started_at.strftime("%d/%m/%Y %H:%M") if run.started_at else "—"),
        ("Tolerância monetária",f"R$ {_fmt_dec(run.monetary_tolerance)}"),
        ("Total de achados",   str(run.total_findings)),
    ]

    for i, (label, value) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    # Cards de resumo
    summary_row = len(meta) + 3
    ws.cell(row=summary_row, column=1, value="Resumo por Severidade").font = Font(bold=True, size=11)
    summary_row += 1

    card_data = [
        ("Críticos",      run.critical_count,   "FFCCCC"),
        ("Alertas",       run.alert_count,       "FFF2CC"),
        ("Divergências",  run.monetary_count,    "CCE5FF"),
        ("Observações",   run.observation_count, "F0F0F0"),
    ]

    for col_offset, (label, count, color) in enumerate(card_data):
        col = col_offset * 2 + 1
        label_cell = ws.cell(row=summary_row, column=col, value=label)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill("solid", fgColor=color)
        label_cell.alignment = Alignment(horizontal="center")

        count_cell = ws.cell(row=summary_row + 1, column=col, value=count)
        count_cell.font = Font(bold=True, size=18)
        count_cell.fill = PatternFill("solid", fgColor=color)
        count_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[summary_row + 1].height = 32

    # Breakdown por registro
    breakdown_row = summary_row + 4
    ws.cell(row=breakdown_row, column=1, value="Achados por Registro").font = Font(bold=True, size=11)
    breakdown_row += 1

    reg_counts: dict[str, int] = {}
    for f in findings:
        reg = f.register_code or "Geral"
        reg_counts[reg] = reg_counts.get(reg, 0) + 1

    _col_header(ws, breakdown_row, ["Registro", "Qtd. Achados"])
    breakdown_row += 1
    for reg, count in sorted(reg_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=breakdown_row, column=1, value=reg)
        ws.cell(row=breakdown_row, column=2, value=count).alignment = Alignment(horizontal="center")
        breakdown_row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def _build_findings_sheet(ws, findings: list[ValidationFinding]) -> None:
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:K1")
    ws["A1"].value = "Achados da Conferência"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = [
        "Severidade", "Tipo", "Regra", "Título", "Descrição",
        "Registro", "Campo", "CFOP", "Valor EFD (R$)",
        "Valor Referência (R$)", "Diferença (R$)", "Status",
    ]
    _col_header(ws, 2, headers)
    ws.row_dimensions[2].height = 18

    for row_idx, f in enumerate(findings, start=3):
        sev_color = SEVERITY_COLORS.get(f.severity, "FFFFFF")
        row_data = [
            SEVERITY_LABELS.get(f.severity, f.severity),
            FINDING_TYPE_LABELS.get(f.finding_type, f.finding_type),
            f.rule_code,
            f.title,
            f.description or "",
            f.register_code or "",
            f.field_name or "",
            f.cfop or "",
            _fmt_dec(f.efd_value),
            _fmt_dec(f.reference_value),
            _fmt_dec(f.difference_value),
            STATUS_LABELS.get(f.status, f.status),
        ]
        fill = PatternFill("solid", fgColor=sev_color)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Larguras fixas
    col_widths = [20, 24, 22, 50, 60, 12, 20, 8, 18, 18, 14, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_suggestions_sheet(ws, suggestions: list[CorrectionSuggestion]) -> None:
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Sugestões de Correção"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = [
        "Status", "Risco", "Linha TXT", "Registro", "Campo",
        "Valor Atual (TXT)", "Valor Sugerido", "Motivo", "Aprovado por", "Aprovado em",
    ]
    _col_header(ws, 2, headers)

    STATUS_SUG = {
        "pending":  ("Pendente",  "FFF2CC"),
        "approved": ("Aprovado",  "CCFFCC"),
        "rejected": ("Rejeitado", "FFCCCC"),
        "applied":  ("Aplicado",  "CCE5FF"),
    }
    RISK_COLORS = {"high": "FFCCCC", "medium": "FFF2CC", "low": "CCFFCC"}

    for row_idx, s in enumerate(suggestions, start=3):
        status_label, status_color = STATUS_SUG.get(s.status, (s.status, "FFFFFF"))
        row_data = [
            status_label,
            {"high": "Alto", "medium": "Médio", "low": "Baixo"}.get(s.risk_level, s.risk_level),
            s.line_number,
            s.register_code,
            s.field_name,
            s.original_value or "",
            s.suggested_value,
            s.suggestion_reason or "",
            s.approved_by or "",
            s.approved_at.strftime("%d/%m/%Y %H:%M") if s.approved_at else "",
        ]
        fill = PatternFill("solid", fgColor=status_color)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 1:
                cell.fill = fill
            elif col_idx == 2:
                cell.fill = PatternFill("solid", fgColor=RISK_COLORS.get(s.risk_level, "FFFFFF"))
            else:
                cell.fill = PatternFill("solid", fgColor="F9F9F9")

    col_widths = [12, 10, 10, 10, 22, 20, 20, 60, 16, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
