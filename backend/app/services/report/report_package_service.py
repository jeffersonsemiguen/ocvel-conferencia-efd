"""
Serviço de geração de pacote de relatórios (ZIP) da competência.
Gera:
  - manifest.json
  - relatorios/resumo_executivo.xlsx
  - arquivos/efd_original.txt (opcional)
  - arquivos/efd_corrigido.txt (opcional)
"""
from __future__ import annotations

import hashlib
import json
import os
import tempfile
import uuid
import zipfile
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.config import settings
from app.models.company import Company
from app.models.efd_file import EfdFile
from app.models.fiscal_period import FiscalPeriod
from app.models.correction import CorrectionSuggestion, CorrectedFile
from app.models.validation import ValidationRun, ValidationFinding
from app.models.period_analytics import ReportPackage
from app.services.consolidation.fiscal_period_dashboard_service import get_period_dashboard


MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

SEVERITY_LABELS = {
    "critico": "Crítico",
    "alerta": "Alerta",
    "divergencia_monetaria": "Divergência Monetária",
    "observacao": "Observação",
}

SEVERITY_COLORS = {
    "critico": "FFCCCC",
    "alerta": "FFF2CC",
    "divergencia_monetaria": "CCE5FF",
    "observacao": "F0F0F0",
}


def _header_style(ws, row: int, col: int, value: str, bg: str = "1F497D") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_resumo_xlsx(db: Session, fiscal_period_id: uuid.UUID, dashboard: dict) -> bytes:
    """Gera bytes do XLSX de resumo executivo."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    # ── Aba Resumo ──────────────────────────────────────────────────────────
    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.column_dimensions["A"].width = 30
    ws_resumo.column_dimensions["B"].width = 40

    period_info = dashboard.get("period", {})
    company_info = dashboard.get("company", {})
    risk_info = dashboard.get("risk", {})
    findings_info = dashboard.get("findings", {})
    suggestions_info = dashboard.get("suggestions", {})

    month_idx = period_info.get("month", 1)
    month_name = MESES[month_idx - 1] if 1 <= month_idx <= 12 else str(month_idx)
    competencia = f"{month_name}/{period_info.get('year', '')}"

    rows = [
        ("Empresa", company_info.get("name", "")),
        ("CNPJ", company_info.get("cnpj", "")),
        ("Competência", competencia),
        ("Status", period_info.get("status", "")),
        ("Score de Risco", risk_info.get("score", 0)),
        ("Nível de Risco", risk_info.get("risk_level", "")),
        ("Total de Achados", findings_info.get("total", 0)),
        ("Achados Críticos", findings_info.get("critical_count", 0)),
        ("Alertas", findings_info.get("warning_count", 0)),
        ("Sugestões Pendentes", suggestions_info.get("pending", 0)),
        ("Sugestões Aprovadas", suggestions_info.get("approved", 0)),
        ("Sugestões Aplicadas", suggestions_info.get("applied", 0)),
        ("Próxima Ação", dashboard.get("next_action", "")),
        ("Gerado em", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")),
    ]

    _header_style(ws_resumo, 1, 1, "Campo")
    _header_style(ws_resumo, 1, 2, "Valor")
    for i, (k, v) in enumerate(rows, start=2):
        ws_resumo.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_resumo.cell(row=i, column=2, value=str(v))

    # ── Aba Principais Achados ──────────────────────────────────────────────
    ws_findings = wb.create_sheet("Principais Achados")
    ws_findings.column_dimensions["A"].width = 20
    ws_findings.column_dimensions["B"].width = 15
    ws_findings.column_dimensions["C"].width = 50
    ws_findings.column_dimensions["D"].width = 15

    headers = ["Regra", "Severidade", "Título", "Registro"]
    for col, h in enumerate(headers, start=1):
        _header_style(ws_findings, 1, col, h)

    # Buscar findings do último run
    last_run = (
        db.query(ValidationRun)
        .filter(
            ValidationRun.fiscal_period_id == fiscal_period_id,
            ValidationRun.status == "completed",
        )
        .order_by(ValidationRun.created_at.desc())
        .first()
    )
    if last_run:
        findings = (
            db.query(ValidationFinding)
            .filter(ValidationFinding.validation_run_id == last_run.id)
            .order_by(ValidationFinding.severity)
            .limit(20)
            .all()
        )
        for i, f in enumerate(findings, start=2):
            color = SEVERITY_COLORS.get(f.severity, "FFFFFF")
            fill = PatternFill(fill_type="solid", fgColor=color)
            for col, val in enumerate([
                f.rule_code or "",
                SEVERITY_LABELS.get(f.severity, f.severity),
                f.title,
                f.register_code or "",
            ], start=1):
                cell = ws_findings.cell(row=i, column=col, value=val)
                cell.fill = fill

    # ── Aba Próxima Ação ────────────────────────────────────────────────────
    ws_action = wb.create_sheet("Proxima Acao")
    ws_action.column_dimensions["A"].width = 60

    ws_action["A1"] = "Recomendação"
    ws_action["A1"].font = Font(bold=True, size=14)
    ws_action["A2"] = dashboard.get("next_action", "")
    ws_action["A2"].alignment = Alignment(wrap_text=True)

    ws_action["A4"] = "Checklist de Pendências"
    ws_action["A4"].font = Font(bold=True)

    checklist = []
    if not findings_info.get("last_run_id"):
        checklist.append("[ ] Executar conferências fiscais")
    if findings_info.get("critical_count", 0) > 0:
        checklist.append(f"[ ] Resolver {findings_info['critical_count']} achado(s) crítico(s)")
    if suggestions_info.get("pending", 0) > 0:
        checklist.append(f"[ ] Revisar {suggestions_info['pending']} sugestão(ões) pendente(s)")
    if suggestions_info.get("approved", 0) > 0 and not dashboard.get("files", {}).get("corrected_count"):
        checklist.append("[ ] Gerar TXT corrigido")
    if dashboard.get("files", {}).get("corrected_count", 0) > 0:
        checklist.append("[x] TXT corrigido gerado — validar no PVA")
    if not checklist:
        checklist.append("[x] Sem pendências críticas")

    for i, item in enumerate(checklist, start=5):
        ws_action.cell(row=i, column=1, value=item)

    # Serializar para bytes
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_report_package(
    db: Session,
    fiscal_period_id: uuid.UUID,
    user_id: Optional[uuid.UUID] = None,
    options: Optional[dict] = None,
) -> ReportPackage:
    """
    Gera ZIP com manifest, resumo_executivo.xlsx e arquivos opcionais.
    Salva em {upload_dir}/packages/{fiscal_period_id}_{timestamp}.zip
    Retorna ReportPackage salvo no banco.
    """
    if options is None:
        options = {}

    period = db.query(FiscalPeriod).filter(FiscalPeriod.id == fiscal_period_id).first()
    if not period:
        raise ValueError(f"Competência {fiscal_period_id} não encontrada")

    company = db.query(Company).filter(Company.id == period.company_id).first()

    # Montar dashboard para dados do relatório
    dashboard = get_period_dashboard(db, fiscal_period_id)

    # Diretório de saída
    packages_dir = os.path.join(settings.upload_dir, "packages")
    os.makedirs(packages_dir, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    package_filename = f"relatorio_{fiscal_period_id}_{timestamp}.zip"
    zip_path = os.path.join(packages_dir, package_filename)

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # manifest.json
        manifest = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "fiscal_period_id": str(fiscal_period_id),
            "company": {
                "id": str(company.id) if company else None,
                "name": company.name if company else None,
                "cnpj": company.cnpj if company else None,
            },
            "period": {
                "year": period.year,
                "month": period.month,
            },
            "contents": [],
        }

        # Resumo executivo XLSX
        xlsx_bytes = _build_resumo_xlsx(db, fiscal_period_id, dashboard)
        zf.writestr("relatorios/resumo_executivo.xlsx", xlsx_bytes)
        manifest["contents"].append("relatorios/resumo_executivo.xlsx")

        # EFD original (opcional)
        if options.get("include_original_efd", False):
            latest_efd = (
                db.query(EfdFile)
                .filter(EfdFile.fiscal_period_id == fiscal_period_id)
                .order_by(EfdFile.created_at.desc())
                .first()
            )
            if latest_efd and hasattr(latest_efd, "stored_path") and latest_efd.stored_path:
                efd_path = latest_efd.stored_path
                if os.path.exists(efd_path):
                    with open(efd_path, "rb") as f:
                        zf.writestr("arquivos/efd_original.txt", f.read())
                    manifest["contents"].append("arquivos/efd_original.txt")

        # EFD corrigido (opcional)
        if options.get("include_corrected_efd", False):
            corrected = (
                db.query(CorrectedFile)
                .filter(CorrectedFile.fiscal_period_id == fiscal_period_id)
                .order_by(CorrectedFile.generated_at.desc())
                .first()
            )
            if corrected and corrected.storage_path and os.path.exists(corrected.storage_path):
                with open(corrected.storage_path, "rb") as f:
                    zf.writestr("arquivos/efd_corrigido.txt", f.read())
                manifest["contents"].append("arquivos/efd_corrigido.txt")

        # Salvar manifest
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    # Calcular hash SHA-256 do ZIP
    sha256 = hashlib.sha256()
    with open(zip_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            sha256.update(chunk)
    file_hash = sha256.hexdigest()
    total_bytes = os.path.getsize(zip_path)

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    pkg = ReportPackage(
        company_id=period.company_id,
        fiscal_period_id=fiscal_period_id,
        package_filename=package_filename,
        storage_path=zip_path,
        file_hash=file_hash,
        total_bytes=total_bytes,
        generated_by=user_id,
        generated_at=now,
        status="generated",
    )
    db.add(pkg)
    db.flush()
    return pkg
