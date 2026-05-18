"""
Importa valores de apuração a partir de XLSX ou CSV.
Colunas obrigatórias: operation_type, tax_type.
"""
from __future__ import annotations

import csv
import io
import uuid
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

import openpyxl

from sqlalchemy.orm import Session

from app.models.apuracao_reference import ApuracaoReferenceValue

VALID_OPERATION_TYPES = {
    "entrada", "saida", "apuracao_icms", "apuracao_icms_st",
    "apuracao_ipi", "ajuste_icms", "ajuste_ipi",
}
VALID_TAX_TYPES = {"icms", "icms_st", "ipi", "difal", "fecop", "outros"}
REQUIRED_COLS = {"operation_type", "tax_type"}
MONEY_COLS = {
    "accounting_value", "icms_base", "icms_amount",
    "icms_st_base", "icms_st_amount", "ipi_base", "ipi_amount", "aliquot",
}


@dataclass
class ImportResult:
    rows_imported: int
    rows_skipped: int
    errors: list[str]


def _to_decimal(value: str | None) -> Decimal | None:
    if not value:
        return None
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except InvalidOperation:
        return None


def _parse_rows(rows: list[dict]) -> tuple[list[dict], list[str]]:
    valid, errors = [], []
    for i, row in enumerate(rows, start=2):
        op = str(row.get("operation_type", "")).strip().lower()
        tax = str(row.get("tax_type", "")).strip().lower()
        if op not in VALID_OPERATION_TYPES:
            errors.append(f"Linha {i}: operation_type inválido '{op}'")
            continue
        if tax not in VALID_TAX_TYPES:
            errors.append(f"Linha {i}: tax_type inválido '{tax}'")
            continue
        row["operation_type"] = op
        row["tax_type"] = tax
        valid.append(row)
    return valid, errors


def import_xlsx(
    db: Session,
    content: bytes,
    company_id: uuid.UUID,
    fiscal_period_id: uuid.UUID,
    pdf_file_id: uuid.UUID | None = None,
) -> ImportResult:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    sheet = wb["apuracao"] if "apuracao" in wb.sheetnames else wb.active
    rows_raw = list(sheet.iter_rows(values_only=True))
    if not rows_raw:
        return ImportResult(0, 0, ["Planilha vazia"])

    headers = [str(h).strip().lower() if h else "" for h in rows_raw[0]]
    missing = REQUIRED_COLS - set(headers)
    if missing:
        return ImportResult(0, 0, [f"Colunas obrigatórias ausentes: {', '.join(missing)}"])

    rows = [dict(zip(headers, row)) for row in rows_raw[1:] if any(c is not None for c in row)]
    return _persist(db, rows, company_id, fiscal_period_id, pdf_file_id, "spreadsheet")


def import_csv(
    db: Session,
    content: bytes,
    company_id: uuid.UUID,
    fiscal_period_id: uuid.UUID,
    pdf_file_id: uuid.UUID | None = None,
) -> ImportResult:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = {h.strip().lower() for h in (reader.fieldnames or [])}
    missing = REQUIRED_COLS - headers
    if missing:
        return ImportResult(0, 0, [f"Colunas obrigatórias ausentes: {', '.join(missing)}"])

    rows = [{k.strip().lower(): v for k, v in row.items()} for row in reader]
    return _persist(db, rows, company_id, fiscal_period_id, pdf_file_id, "spreadsheet")


def _persist(
    db: Session,
    rows: list[dict],
    company_id: uuid.UUID,
    fiscal_period_id: uuid.UUID,
    pdf_file_id: uuid.UUID | None,
    source_type: str,
) -> ImportResult:
    valid_rows, errors = _parse_rows(rows)
    imported = 0

    for i, row in enumerate(valid_rows):
        db.add(ApuracaoReferenceValue(
            company_id=company_id,
            fiscal_period_id=fiscal_period_id,
            pdf_file_id=pdf_file_id,
            source_type=source_type,
            source_label=str(row.get("source_label", "") or "").strip() or None,
            operation_type=row["operation_type"],
            tax_type=row["tax_type"],
            cfop=str(row.get("cfop", "") or "").strip() or None,
            cst=str(row.get("cst", "") or "").strip() or None,
            csosn=str(row.get("csosn", "") or "").strip() or None,
            cst_ipi=str(row.get("cst_ipi", "") or "").strip() or None,
            aliquot=_to_decimal(str(row.get("aliquot") or "")),
            accounting_value=_to_decimal(str(row.get("accounting_value") or "")),
            icms_base=_to_decimal(str(row.get("icms_base") or "")),
            icms_amount=_to_decimal(str(row.get("icms_amount") or "")),
            icms_st_base=_to_decimal(str(row.get("icms_st_base") or "")),
            icms_st_amount=_to_decimal(str(row.get("icms_st_amount") or "")),
            ipi_base=_to_decimal(str(row.get("ipi_base") or "")),
            ipi_amount=_to_decimal(str(row.get("ipi_amount") or "")),
            adjustment_code=str(row.get("adjustment_code", "") or "").strip() or None,
            adjustment_description=str(row.get("adjustment_description", "") or "").strip() or None,
            source_row=i + 2,
            is_reviewed=False,
        ))
        imported += 1

    db.flush()
    return ImportResult(rows_imported=imported, rows_skipped=len(rows) - len(valid_rows), errors=errors)
