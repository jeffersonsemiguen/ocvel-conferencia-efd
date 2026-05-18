import hashlib
import os
import uuid
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user
from app.models.pr_adjustment import (
    PrAdjustmentCode,
    PrAdjustmentImportBatch,
    PrAdjustmentValidationResult,
)

router = APIRouter(
    dependencies=[Depends(get_current_user)],
    prefix="/api/v1/pr-adjustment-codes",
    tags=["pr-adjustment"],
)

SEED_FILE_ENV = "PR_ADJUSTMENT_MD_PATH"


# ── Pydantic schemas ──────────────────────────────────────────────────────────

class CreatePrCode(BaseModel):
    code: str
    table_type: str = "ajuste_apuracao"
    description: str
    short_description: Optional[str] = None
    register_expected: Optional[str] = None
    apuracao_type: Optional[str] = None
    adjustment_nature: Optional[str] = None
    operation_scope: Optional[str] = None
    requires_e112: bool = False
    requires_e113: bool = False
    optional_e112: bool = False
    optional_e113: bool = False
    requires_fiscal_document: bool = False
    requires_process: bool = False
    requires_auxiliary_ie: bool = False
    requires_item: bool = False
    requires_participant: bool = False
    valid_from: Optional[date] = None
    valid_to: Optional[date] = None
    orientation_text: Optional[str] = None
    source_name: Optional[str] = None
    source_version: Optional[str] = None
    is_active: bool = True


class PatchPrCode(BaseModel):
    description: Optional[str] = None
    short_description: Optional[str] = None
    table_type: Optional[str] = None
    register_expected: Optional[str] = None
    apuracao_type: Optional[str] = None
    adjustment_nature: Optional[str] = None
    requires_e112: Optional[bool] = None
    requires_e113: Optional[bool] = None
    optional_e112: Optional[bool] = None
    optional_e113: Optional[bool] = None
    requires_fiscal_document: Optional[bool] = None
    requires_process: Optional[bool] = None
    requires_auxiliary_ie: Optional[bool] = None
    requires_item: Optional[bool] = None
    requires_participant: Optional[bool] = None
    valid_from: Optional[date] = None
    valid_to: Optional[date] = None
    orientation_text: Optional[str] = None
    source_name: Optional[str] = None
    source_version: Optional[str] = None
    is_active: Optional[bool] = None


# ── Seed endpoints (existing) ─────────────────────────────────────────────────

@router.post("/seed", status_code=status.HTTP_201_CREATED)
def seed_from_file(db: Session = Depends(get_db)):
    """
    Carrega os códigos de ajuste PR a partir do arquivo .md configurado
    na variável de ambiente PR_ADJUSTMENT_MD_PATH.
    """
    md_path = os.environ.get(SEED_FILE_ENV)
    if not md_path or not os.path.exists(md_path):
        raise HTTPException(
            422,
            f"Arquivo não encontrado. Configure {SEED_FILE_ENV} com o caminho do .md.",
        )
    return _do_seed(db, md_path)


@router.post("/seed-upload", status_code=status.HTTP_201_CREATED)
def seed_from_upload(file: UploadFile, db: Session = Depends(get_db)):
    """Upload direto do arquivo .md da Tabela 5.1.1."""
    if not (file.filename or "").lower().endswith(".md"):
        raise HTTPException(400, "Apenas arquivos .md são aceitos")

    import tempfile
    content = file.file.read()
    with tempfile.NamedTemporaryFile(suffix=".md", delete=False, mode="wb") as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = _do_seed(db, tmp_path)
    finally:
        os.unlink(tmp_path)

    return result


# ── XLSX Import endpoint (Sprint 5) ─────────────────────────────────────────

@router.post("/import", status_code=status.HTTP_201_CREATED)
def import_from_xlsx(
    file: UploadFile,
    source_name: Optional[str] = Query(None),
    source_version: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Importa códigos de ajuste PR a partir de arquivo XLSX.
    Colunas esperadas: code, table_type, description, short_description,
    register_expected, apuracao_type, adjustment_nature, requires_e112,
    requires_e113, requires_fiscal_document, requires_process, valid_from,
    valid_to, orientation_text, source_name, source_version, is_active
    """
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "Apenas arquivos .xlsx são aceitos")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl não instalado. Execute: uv add openpyxl")

    content = file.file.read()
    file_hash = hashlib.sha256(content).hexdigest()

    # Create import batch
    batch = PrAdjustmentImportBatch(
        original_filename=file.filename or "unknown.xlsx",
        file_hash=file_hash,
        source_name=source_name,
        source_version=source_version,
        status="processing",
        imported_at=datetime.now(timezone.utc),
    )
    db.add(batch)
    db.flush()

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        batch.status = "failed"
        batch.error_summary = str(exc)
        db.commit()
        raise HTTPException(422, f"Erro ao abrir XLSX: {exc}")

    ws = wb.active
    headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def _col(row, name: str):
        try:
            idx = headers.index(name)
            return row[idx].value
        except (ValueError, IndexError):
            return None

    def _bool_val(v) -> bool:
        if v is None:
            return False
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in ("1", "true", "sim", "yes", "s")

    def _date_val(v) -> Optional[date]:
        if v is None:
            return None
        if isinstance(v, (datetime,)):
            return v.date()
        if isinstance(v, date):
            return v
        try:
            return date.fromisoformat(str(v).strip())
        except Exception:
            return None

    total = 0
    imported = 0
    failed = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        code = _col(row, "code")
        if code is None or str(code).strip() == "":
            continue
        total += 1
        code = str(code).strip().upper()
        ttype = str(_col(row, "table_type") or "ajuste_apuracao").strip()

        try:
            # Try to find existing by code + table_type + valid_from
            vf = _date_val(_col(row, "valid_from"))
            existing = (
                db.query(PrAdjustmentCode)
                .filter(
                    PrAdjustmentCode.code == code,
                    PrAdjustmentCode.table_type == ttype,
                )
                .first()
            )

            fields = dict(
                description=str(_col(row, "description") or code),
                short_description=str(_col(row, "short_description") or "") or None,
                register_expected=str(_col(row, "register_expected") or "") or None,
                apuracao_type=str(_col(row, "apuracao_type") or "") or None,
                adjustment_nature=str(_col(row, "adjustment_nature") or "") or None,
                requires_e112=_bool_val(_col(row, "requires_e112")),
                requires_e113=_bool_val(_col(row, "requires_e113")),
                requires_fiscal_document=_bool_val(_col(row, "requires_fiscal_document")),
                requires_process=_bool_val(_col(row, "requires_process")),
                valid_from=vf,
                valid_to=_date_val(_col(row, "valid_to")),
                orientation_text=str(_col(row, "orientation_text") or "") or None,
                source_name=source_name or str(_col(row, "source_name") or "") or None,
                source_version=source_version or str(_col(row, "source_version") or "") or None,
                is_active=_bool_val(_col(row, "is_active") if _col(row, "is_active") is not None else True),
                import_batch_id=batch.id,
            )

            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
            else:
                db.add(PrAdjustmentCode(code=code, table_type=ttype, **fields))

            imported += 1
        except Exception as exc:
            failed += 1
            errors.append(f"Linha {row_num}: {exc}")

    batch.records_total = total
    batch.records_imported = imported
    batch.records_failed = failed
    batch.status = "imported" if failed == 0 else "imported_with_errors"
    if errors:
        batch.error_summary = "\n".join(errors[:50])

    db.commit()

    return {
        "batch_id": str(batch.id),
        "status": batch.status,
        "records_total": total,
        "records_imported": imported,
        "records_failed": failed,
        "errors": errors[:10] if errors else [],
    }


# ── List / Get / Create / Patch ──────────────────────────────────────────────

@router.get("/")
def list_codes(
    code: Optional[str] = Query(None),
    table_type: Optional[str] = Query(None),
    valid_on: Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(PrAdjustmentCode).filter(PrAdjustmentCode.is_active == True)
    if code:
        q = q.filter(PrAdjustmentCode.code.ilike(f"{code}%"))
    if table_type:
        q = q.filter(PrAdjustmentCode.table_type == table_type)
    if valid_on:
        q = q.filter(
            (PrAdjustmentCode.valid_from == None) | (PrAdjustmentCode.valid_from <= valid_on),
            (PrAdjustmentCode.valid_to == None) | (PrAdjustmentCode.valid_to >= valid_on),
        )
    codes = q.order_by(PrAdjustmentCode.code).all()
    return [_to_dict(c) for c in codes]


@router.get("/{code_str}")
def get_code(code_str: str, db: Session = Depends(get_db)):
    c = db.query(PrAdjustmentCode).filter(PrAdjustmentCode.code == code_str.upper()).first()
    if not c:
        raise HTTPException(404, f"Código {code_str} não encontrado")
    return _to_dict(c)


@router.post("/", status_code=status.HTTP_201_CREATED)
def create_code(body: CreatePrCode, db: Session = Depends(get_db)):
    obj = PrAdjustmentCode(**body.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return _to_dict(obj)


@router.patch("/{code_id}", status_code=status.HTTP_200_OK)
def patch_code(code_id: uuid.UUID, body: PatchPrCode, db: Session = Depends(get_db)):
    obj = db.query(PrAdjustmentCode).filter(PrAdjustmentCode.id == code_id).first()
    if not obj:
        raise HTTPException(404, "Código não encontrado")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return _to_dict(obj)


# ── Validation results endpoint ──────────────────────────────────────────────

@router.get("/validation-results/{run_id}", tags=["pr-adjustment", "validation"])
def get_pr_validation_results(run_id: uuid.UUID, db: Session = Depends(get_db)):
    results = (
        db.query(PrAdjustmentValidationResult)
        .filter(PrAdjustmentValidationResult.validation_run_id == run_id)
        .order_by(PrAdjustmentValidationResult.line_number)
        .all()
    )
    return [_result_to_dict(r) for r in results]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _do_seed(db: Session, md_path: str) -> dict:
    from app.services.pr_adjustment.md_parser import parse_markdown
    parsed = parse_markdown(md_path)

    inserted = updated = 0
    for p in parsed:
        existing = db.query(PrAdjustmentCode).filter(PrAdjustmentCode.code == p.code).first()
        if existing:
            existing.description = p.description
            existing.start_date = p.start_date
            existing.end_date = p.end_date
            existing.adjustment_text = p.adjustment_text
            existing.requires_e112 = p.requires_e112
            existing.requires_e113 = p.requires_e113
            existing.optional_e112 = p.optional_e112
            existing.optional_e113 = p.optional_e113
            existing.page_ref = p.page_ref
            existing.is_active = True
            updated += 1
        else:
            db.add(PrAdjustmentCode(
                code=p.code,
                description=p.description,
                start_date=p.start_date,
                end_date=p.end_date,
                adjustment_text=p.adjustment_text,
                requires_e112=p.requires_e112,
                requires_e113=p.requires_e113,
                optional_e112=p.optional_e112,
                optional_e113=p.optional_e113,
                page_ref=p.page_ref,
                is_active=True,
            ))
            inserted += 1

    db.commit()
    return {"inserted": inserted, "updated": updated, "total": inserted + updated}


def _to_dict(c: PrAdjustmentCode) -> dict:
    return {
        "id": str(c.id),
        "code": c.code,
        "table_type": c.table_type,
        "description": c.description,
        "short_description": c.short_description,
        "register_expected": c.register_expected,
        "apuracao_type": c.apuracao_type,
        "adjustment_nature": c.adjustment_nature,
        "start_date": c.start_date,
        "end_date": c.end_date,
        "adjustment_text": c.adjustment_text,
        "requires_e112": c.requires_e112,
        "requires_e113": c.requires_e113,
        "optional_e112": c.optional_e112,
        "optional_e113": c.optional_e113,
        "requires_fiscal_document": c.requires_fiscal_document,
        "requires_process": c.requires_process,
        "requires_auxiliary_ie": c.requires_auxiliary_ie,
        "requires_item": c.requires_item,
        "requires_participant": c.requires_participant,
        "valid_from": c.valid_from.isoformat() if c.valid_from else None,
        "valid_to": c.valid_to.isoformat() if c.valid_to else None,
        "orientation_text": c.orientation_text,
        "source_name": c.source_name,
        "source_version": c.source_version,
        "page_ref": c.page_ref,
        "is_active": c.is_active,
        "import_batch_id": str(c.import_batch_id) if c.import_batch_id else None,
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
    }


def _result_to_dict(r: PrAdjustmentValidationResult) -> dict:
    return {
        "id": str(r.id),
        "validation_run_id": str(r.validation_run_id),
        "efd_file_id": str(r.efd_file_id),
        "register_code": r.register_code,
        "line_number": r.line_number,
        "adjustment_code": r.adjustment_code,
        "adjustment_table_type": r.adjustment_table_type,
        "pr_adjustment_code_id": str(r.pr_adjustment_code_id) if r.pr_adjustment_code_id else None,
        "validation_rule_code": r.validation_rule_code,
        "status": r.status,
        "severity": r.severity,
        "message": r.message,
        "requires_e112": r.requires_e112,
        "has_e112": r.has_e112,
        "requires_e113": r.requires_e113,
        "has_e113": r.has_e113,
        "requires_process": r.requires_process,
        "has_process": r.has_process,
        "requires_fiscal_document": r.requires_fiscal_document,
        "has_fiscal_document": r.has_fiscal_document,
        "requires_auxiliary_ie": r.requires_auxiliary_ie,
        "has_auxiliary_ie": r.has_auxiliary_ie,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }
